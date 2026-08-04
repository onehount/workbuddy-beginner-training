# -*- coding: utf-8 -*-
"""
生成 Demo 输入表格（04-demo-inputs/spreadsheets/）

⚠️ 关键教学设计：所有汇总值与转化率都由 Excel 公式计算，不写死数字。
   这样学员打开表格点一下单元格，就能看到口径。

License: MIT
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import case_data as C

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
OUT = os.path.join(ROOT, "04-demo-inputs", "spreadsheets")

# ---------- 样式 ----------
NAVY = "1B3A6B"
BLUE = "2E6BE6"
LIGHT = "EAF1FD"
GREY = "F5F6F8"

F_TITLE = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
F_HEAD = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
F_BODY = Font(name="微软雅黑", size=10)
F_BOLD = Font(name="微软雅黑", size=10, bold=True)
F_NOTE = Font(name="微软雅黑", size=9, color="666666")

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HEAD = PatternFill("solid", fgColor=BLUE)
FILL_SUB = PatternFill("solid", fgColor=LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor="FFE9CC")

THIN = Side(style="thin", color="C9D4E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def set_props(wb):
    p = wb.properties
    p.creator = C.AUTHOR
    p.lastModifiedBy = C.AUTHOR
    p.title = ""
    p.subject = "WorkBuddy 零基础教学项目 · 虚构演示数据"
    p.description = C.DISCLAIMER
    p.keywords = "教学演示;虚构数据;WorkBuddy Training"
    p.category = "Training Material"
    p.revision = "1"


def title_row(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = FILL_TITLE
    c.alignment = CENTER
    ws.row_dimensions[row].height = 30


def note_row(ws, text, ncols, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_NOTE
    c.alignment = LEFT
    ws.row_dimensions[row].height = 18


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 03_参会人员与岗位.xlsx
# ============================================================
def gen_people():
    wb = Workbook()
    set_props(wb)
    ws = wb.active
    ws.title = "参会人员"

    ncols = 5
    title_row(ws, f"{C.BRAND_CN} {C.DEPARTMENT} · 月度复盘会参会人员与岗位", ncols)
    note_row(ws, f"会议：{C.MEETING_TITLE}　｜　时间：{C.MEETING_DATE} {C.MEETING_TIME}　｜　地点：{C.MEETING_PLACE}", ncols, 2)
    note_row(ws, f"⚠️ {C.DISCLAIMER}", ncols, 3)

    header_row(ws, ["序号", "姓名", "岗位", "所属", "本月主要负责"], 5)
    r = 6
    for i, (name, role, dept, duty) in enumerate(C.PEOPLE, start=1):
        vals = [i, name.replace("　", ""), role, dept, duty]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_BODY
            c.border = BORDER
            c.alignment = CENTER if j <= 4 else LEFT
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value="参会人数合计")
    c.font = F_BOLD
    c.fill = FILL_TOTAL
    c.alignment = CENTER
    c.border = BORDER
    c2 = ws.cell(row=r, column=3, value=f"=COUNTA(B6:B{5 + len(C.PEOPLE)})")
    c2.font = F_BOLD
    c2.fill = FILL_TOTAL
    c2.alignment = CENTER
    c2.border = BORDER

    widths(ws, [8, 12, 18, 12, 46])
    ws.freeze_panes = "A6"

    path = os.path.join(OUT, "03_参会人员与岗位.xlsx")
    wb.save(path)
    return path


# ============================================================
# 04_市场活动完成情况.xlsx
# ============================================================
def gen_campaigns():
    wb = Workbook()
    set_props(wb)
    ws = wb.active
    ws.title = "活动完成情况"

    ncols = 9
    title_row(ws, f"{C.BRAND_CN} {C.DEPARTMENT} · {C.PERIOD_LABEL} 市场活动完成情况", ncols)
    note_row(ws, f"统计周期：{C.PERIOD_RANGE}", ncols, 2)
    note_row(ws, f"⚠️ {C.DISCLAIMER}", ncols, 3)

    header_row(ws, ["序号", "活动名称", "开始日期", "实际结束", "计划结束",
                    "状态", "负责人", "协同", "说明"], 5)
    r = 6
    for (idx, name, start, end, plan_end, status, owner, helper, ch, desc) in C.CAMPAIGNS:
        vals = [idx, name, start, end if end else "—", plan_end,
                status, owner.replace("　", ""), helper.replace("　", ""), desc]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_BODY
            c.border = BORDER
            c.alignment = LEFT if j == 9 else CENTER
            if j == 6 and v == "延期":
                c.font = Font(name="微软雅黑", size=10, bold=True, color="C0392B")
        ws.row_dimensions[r].height = 30
        r += 1

    last = r - 1
    r += 1

    # 汇总区（全部用公式）
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value="汇总指标")
    c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = CENTER; c.border = BORDER
    c = ws.cell(row=r, column=3, value="计算值")
    c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = CENTER; c.border = BORDER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    c = ws.cell(row=r, column=4, value="计算公式（点击单元格可查看）")
    c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = CENTER; c.border = BORDER
    r += 1

    summary = [
        ("市场活动数量", f"=COUNTA(B6:B{last})", "COUNTA(活动名称列)"),
        ("已完成活动数量", f'=COUNTIF(F6:F{last},"已完成")', 'COUNTIF(状态列,"已完成")'),
        ("延期活动数量", f'=COUNTIF(F6:F{last},"延期")', 'COUNTIF(状态列,"延期")'),
        ("活动完成率", f'=COUNTIF(F6:F{last},"已完成")/COUNTA(B6:B{last})',
         '已完成数量 ÷ 活动总数'),
    ]
    for label, formula, expl in summary:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=label)
        c.font = F_BOLD; c.fill = FILL_SUB; c.alignment = CENTER; c.border = BORDER
        c = ws.cell(row=r, column=3, value=formula)
        c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = CENTER; c.border = BORDER
        if "完成率" in label:
            c.number_format = "0.0%"
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        c = ws.cell(row=r, column=4, value=expl)
        c.font = F_NOTE; c.alignment = LEFT; c.border = BORDER
        r += 1

    widths(ws, [7, 24, 13, 13, 13, 9, 11, 11, 44])
    ws.freeze_panes = "A6"

    path = os.path.join(OUT, "04_市场活动完成情况.xlsx")
    wb.save(path)
    return path


# ============================================================
# 05_渠道曝光与线索数据.xlsx
# ============================================================
def gen_channels():
    wb = Workbook()
    set_props(wb)

    # ---- Sheet 1: 周度明细 ----
    ws = wb.active
    ws.title = "周度明细"
    ncols = 6

    title_row(ws, f"{C.BRAND_CN} {C.DEPARTMENT} · {C.PERIOD_LABEL} 渠道曝光与线索数据（周度明细）", ncols)
    note_row(ws, C.METRIC_NOTE, ncols, 2)
    note_row(ws, f"⚠️ {C.DISCLAIMER}", ncols, 3)

    header_row(ws, ["渠道", "统计周", "日期范围", "曝光量", "线索数量", "有效线索"], 5)

    r = 6
    week_map = dict(C.WEEKS)
    detail_start = r
    for ch in C.CHANNELS:
        for (wk, exp, lead, qual) in C.CHANNEL_WEEKLY[ch]:
            vals = [ch, wk, week_map[wk], exp, lead, qual]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = F_BODY
                c.border = BORDER
                if j >= 4:
                    c.alignment = RIGHT
                    c.number_format = "#,##0"
                else:
                    c.alignment = CENTER
            r += 1
    detail_end = r - 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value="全渠道合计")
    c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = CENTER; c.border = BORDER
    for j, col in enumerate(["D", "E", "F"], start=4):
        c = ws.cell(row=r, column=j, value=f"=SUM({col}{detail_start}:{col}{detail_end})")
        c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = RIGHT
        c.number_format = "#,##0"; c.border = BORDER

    widths(ws, [16, 10, 18, 14, 13, 13])
    ws.freeze_panes = "A6"

    # ---- Sheet 2: 渠道汇总 ----
    ws2 = wb.create_sheet("渠道汇总")
    ncols2 = 7
    title_row(ws2, f"{C.BRAND_CN} {C.DEPARTMENT} · {C.PERIOD_LABEL} 渠道汇总", ncols2)
    note_row(ws2, C.METRIC_NOTE, ncols2, 2)
    note_row(ws2, "本表所有数值均由「周度明细」工作表通过 SUMIF 公式计算得出，未手工填写。", ncols2, 3)

    header_row(ws2, ["渠道", "曝光量", "线索数量", "有效线索", "转化率",
                     "曝光量占比", "有效线索占比"], 5)

    r2 = 6
    sum_start = r2
    for ch in C.CHANNELS:
        c = ws2.cell(row=r2, column=1, value=ch)
        c.font = F_BODY; c.alignment = CENTER; c.border = BORDER
        for j, col in enumerate(["D", "E", "F"], start=2):
            c = ws2.cell(row=r2, column=j,
                         value=f"=SUMIF(周度明细!$A${detail_start}:$A${detail_end},$A{r2},周度明细!${col}${detail_start}:${col}${detail_end})")
            c.font = F_BODY; c.alignment = RIGHT; c.number_format = "#,##0"; c.border = BORDER
        # 转化率
        c = ws2.cell(row=r2, column=5, value=f"=D{r2}/C{r2}")
        c.font = F_BODY; c.alignment = RIGHT; c.number_format = "0.0%"; c.border = BORDER
        r2 += 1
    sum_end = r2 - 1

    # 合计行
    total_row = r2
    c = ws2.cell(row=total_row, column=1, value="合计")
    c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = CENTER; c.border = BORDER
    for j, col in enumerate(["B", "C", "D"], start=2):
        c = ws2.cell(row=total_row, column=j, value=f"=SUM({col}{sum_start}:{col}{sum_end})")
        c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = RIGHT
        c.number_format = "#,##0"; c.border = BORDER
    c = ws2.cell(row=total_row, column=5, value=f"=D{total_row}/C{total_row}")
    c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = RIGHT
    c.number_format = "0.0%"; c.border = BORDER

    # 占比列（需要总计行，所以最后填）
    for rr in range(sum_start, sum_end + 1):
        c = ws2.cell(row=rr, column=6, value=f"=B{rr}/$B${total_row}")
        c.font = F_BODY; c.alignment = RIGHT; c.number_format = "0.0%"; c.border = BORDER
        c = ws2.cell(row=rr, column=7, value=f"=D{rr}/$D${total_row}")
        c.font = F_BODY; c.alignment = RIGHT; c.number_format = "0.0%"; c.border = BORDER
    for j in (6, 7):
        c = ws2.cell(row=total_row, column=j, value="100.0%")
        c.font = F_BOLD; c.fill = FILL_TOTAL; c.alignment = RIGHT; c.border = BORDER

    # 口径警示
    r2 = total_row + 2
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=ncols2)
    c = ws2.cell(row=r2, column=1,
                 value="⚠️ 注意：整体转化率 = 有效线索合计 ÷ 线索合计，"
                       "不是五个渠道转化率的算术平均值。两者结果不同。")
    c.font = Font(name="微软雅黑", size=9, bold=True, color="C0392B")
    c.alignment = LEFT
    ws2.row_dimensions[r2].height = 20

    r2 += 1
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=ncols2)
    c = ws2.cell(row=r2, column=1,
                 value=f"　　正确算法：=D{total_row}/C{total_row}　"
                       f"　　错误算法：=AVERAGE(E{sum_start}:E{sum_end})　（两者相差约 4.8 个百分点）")
    c.font = F_NOTE
    c.alignment = LEFT

    widths(ws2, [16, 14, 13, 13, 12, 14, 15])
    ws2.freeze_panes = "A6"

    # ---- Sheet 3: 数据字典 ----
    ws3 = wb.create_sheet("数据字典")
    ncols3 = 4
    title_row(ws3, "数据字典与计算口径", ncols3)
    header_row(ws3, ["字段", "定义", "单位", "来源 / 计算方式"], 3)

    DICT = [
        ("渠道", "本月投放或运营的对外触点", "—", "固定 5 个：小红书、抖音、微信公众号、微信社群、线下门店"),
        ("统计周", "自然周划分，W4 含 7/29–7/31 共 10 天", "—", "W1 07-01~07-07；W2 07-08~07-14；W3 07-15~07-21；W4 07-22~07-31"),
        ("曝光量", "内容被展示的次数（去重前）", "次", "各渠道后台导出"),
        ("线索数量", "留下联系方式或完成互动动作的用户数", "条", "各渠道后台导出"),
        ("有效线索", "经人工初筛，符合目标客群且联系方式可用的线索", "条", "人工筛选后统计"),
        ("转化率", "有效线索占全部线索的比例", "%", "有效线索数量 ÷ 线索数量 × 100%"),
        ("整体转化率", "全渠道汇总口径的转化率", "%", "有效线索合计 ÷ 线索合计 × 100%（非各渠道转化率的平均值）"),
        ("曝光量占比", "单渠道曝光量占全渠道曝光量的比例", "%", "单渠道曝光量 ÷ 全渠道曝光量合计"),
        ("有效线索占比", "单渠道有效线索占全渠道有效线索的比例", "%", "单渠道有效线索 ÷ 全渠道有效线索合计"),
    ]
    r3 = 4
    for field, defi, unit, src in DICT:
        for j, v in enumerate([field, defi, unit, src], start=1):
            c = ws3.cell(row=r3, column=j, value=v)
            c.font = F_BOLD if j == 1 else F_BODY
            c.border = BORDER
            c.alignment = CENTER if j in (1, 3) else LEFT
        ws3.row_dimensions[r3].height = 30
        r3 += 1

    r3 += 1
    ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=ncols3)
    c = ws3.cell(row=r3, column=1, value="本表不含预算、成本、客单价、销售额、用户画像数据。任何涉及上述内容的分析结论均无数据支撑。")
    c.font = Font(name="微软雅黑", size=9, bold=True, color="C0392B")
    c.alignment = LEFT

    widths(ws3, [16, 38, 8, 56])

    path = os.path.join(OUT, "05_渠道曝光与线索数据.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    C.verify()
    os.makedirs(OUT, exist_ok=True)
    for fn in (gen_people, gen_campaigns, gen_channels):
        p = fn()
        print(f"[OK] {os.path.relpath(p, ROOT)}")
    print("\n表格生成完成。所有汇总值与转化率均为 Excel 公式，非硬编码。")
